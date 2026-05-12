import io
from django.http import HttpResponse
from django.utils.timezone import localtime
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated
from apps.users.permissions import IsAdminOrOfficial
from apps.beneficiaries.models import Beneficiary, Household, Family
from apps.cycles.models import ProgramCycle, CycleApplication, ParticipationRecord
from apps.audit.models import AuditLog
from apps.criteria.models import Criterion
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

HEADER_FILL = PatternFill(start_color='1a56db', end_color='1a56db', fill_type='solid')
HEADER_FONT = Font(bold=True, color='FFFFFF', size=10)


def _style_headers(ws, row_num):
    for cell in ws[row_num]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal='center', vertical='center')


def _auto_width(ws):
    for col in ws.columns:
        max_len = max((len(str(cell.value or '')) for cell in col), default=8)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 3, 55)


def _excel_response(wb, filename):
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    resp = HttpResponse(
        buf.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    resp['Content-Disposition'] = f'attachment; filename="{filename}"'
    return resp


class CycleRankingReportView(APIView):
    permission_classes = [IsAuthenticated, IsAdminOrOfficial]

    def get(self, request):
        cycle_id = request.query_params.get('cycle_id')
        if not cycle_id:
            return Response({'detail': 'cycle_id is required.'}, status=400)

        try:
            cycle = ProgramCycle.objects.get(pk=cycle_id)
        except ProgramCycle.DoesNotExist:
            return Response({'detail': 'Cycle not found.'}, status=404)

        applications = (
            CycleApplication.objects.filter(cycle=cycle)
            .select_related('beneficiary')
            .prefetch_related('beneficiary__indicators__criterion')
            .order_by('rank_position', '-computed_score')
        )
        criteria = list(Criterion.objects.filter(is_active=True).order_by('name'))

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Ranking'

        total_cols = max(5 + len(criteria), 5)
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_cols)
        ws.cell(1, 1).value = f'Cycle Ranking Report — {cycle.cycle_name}'
        ws.cell(1, 1).font = Font(bold=True, size=12)
        ws.cell(2, 1).value = f'Slots: {cycle.slots}  |  Period: {cycle.start_date} to {cycle.end_date}'
        ws.cell(2, 1).font = Font(italic=True, size=9)

        headers = ['Rank', 'Beneficiary Name', 'Score', 'Status', 'Application Date'] + [c.name for c in criteria]
        ws.append([])
        ws.append(headers)
        _style_headers(ws, 4)

        for app in applications:
            indicator_map = {
                str(ind.criterion_id): round(float(ind.value), 4)
                for ind in app.beneficiary.indicators.all()
            }
            ws.append([
                app.rank_position or '',
                app.beneficiary.full_name,
                round(float(app.computed_score), 4) if app.computed_score else '',
                app.get_status_display(),
                str(app.application_date),
            ] + [indicator_map.get(str(c.id), '') for c in criteria])

        _auto_width(ws)
        safe_name = cycle.cycle_name.replace(' ', '_').replace('/', '-')
        return _excel_response(wb, f'cycle_ranking_{safe_name}.xlsx')


class BeneficiaryMasterlistReportView(APIView):
    permission_classes = [IsAuthenticated, IsAdminOrOfficial]

    def get(self, request):
        beneficiaries = (
            Beneficiary.objects
            .select_related('family__household')
            .order_by('full_name')
        )

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Beneficiary Masterlist'

        ws.merge_cells('A1:N1')
        ws.cell(1, 1).value = 'Beneficiary Masterlist'
        ws.cell(1, 1).font = Font(bold=True, size=12)

        headers = [
            'Full Name', 'Age', 'Gender', 'Civil Status', 'Role in Household',
            'Sectors', 'Employment Status', 'Monthly Income',
            'Household Size', 'Dependents', 'Housing Condition',
            'TUPAD Eligible', 'Household Code', 'Purok / Address',
        ]
        ws.append([])
        ws.append(headers)
        _style_headers(ws, 3)

        for b in beneficiaries:
            household = b.family.household if b.family_id else None
            ws.append([
                b.full_name,
                b.age,
                b.get_gender_display(),
                b.get_civil_status_display(),
                b.get_role_display(),
                ', '.join(b.sectors) if b.sectors else '',
                b.get_employment_status_display(),
                float(b.monthly_income),
                b.household_size,
                b.num_dependents,
                b.get_housing_condition_display(),
                'Yes' if b.is_tupad_eligible else 'No',
                household.household_code if household else '',
                ((household.purok + ' — ' if household and household.purok else '') + (household.address if household else '')).strip(),
            ])

        _auto_width(ws)
        return _excel_response(wb, 'beneficiary_masterlist.xlsx')


class ParticipationHistoryReportView(APIView):
    permission_classes = [IsAuthenticated, IsAdminOrOfficial]

    def get(self, request):
        cycle_id = request.query_params.get('cycle_id')
        qs = (
            ParticipationRecord.objects
            .select_related('beneficiary', 'cycle', 'recorded_by')
            .order_by('cycle__start_date', 'beneficiary__full_name')
        )
        if cycle_id:
            qs = qs.filter(cycle_id=cycle_id)

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Participation History'

        ws.merge_cells('A1:G1')
        ws.cell(1, 1).value = 'Participation History Report'
        ws.cell(1, 1).font = Font(bold=True, size=12)

        headers = ['Beneficiary Name', 'Cycle', 'Project', 'Days Worked', 'Start Date', 'End Date', 'Recorded By']
        ws.append([])
        ws.append(headers)
        _style_headers(ws, 3)

        for rec in qs:
            recorder = rec.recorded_by
            recorder_name = getattr(recorder, 'full_name', None) or str(recorder)
            ws.append([
                rec.beneficiary.full_name,
                rec.cycle.cycle_name,
                rec.project_name,
                rec.days_worked,
                str(rec.participation_start),
                str(rec.participation_end),
                recorder_name,
            ])

        _auto_width(ws)
        return _excel_response(wb, 'participation_history.xlsx')


class AuditTrailReportView(APIView):
    permission_classes = [IsAuthenticated, IsAdminOrOfficial]

    def get(self, request):
        start = request.query_params.get('start_date')
        end = request.query_params.get('end_date')

        qs = AuditLog.objects.select_related('user').order_by('-timestamp')
        if start:
            qs = qs.filter(timestamp__date__gte=start)
        if end:
            qs = qs.filter(timestamp__date__lte=end)

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Audit Trail'

        ws.merge_cells('A1:F1')
        ws.cell(1, 1).value = 'Audit Trail Report'
        ws.cell(1, 1).font = Font(bold=True, size=12)

        headers = ['Timestamp', 'User', 'Action', 'Target Table', 'Target ID', 'Details']
        ws.append([])
        ws.append(headers)
        _style_headers(ws, 3)

        for log in qs:
            ws.append([
                localtime(log.timestamp).strftime('%Y-%m-%d %H:%M:%S'),
                getattr(log.user, 'full_name', None) or str(log.user) if log.user else 'System',
                log.action,
                log.target_table,
                log.target_id or '',
                str(log.details) if log.details else '',
            ])

        _auto_width(ws)
        return _excel_response(wb, 'audit_trail.xlsx')


class HouseholdReportView(APIView):
    permission_classes = [IsAuthenticated, IsAdminOrOfficial]

    def get(self, request):
        households = (
            Household.objects
            .prefetch_related('families__members')
            .order_by('household_code')
        )

        wb = openpyxl.Workbook()

        # Sheet 1 — Households
        ws1 = wb.active
        ws1.title = 'Households'
        ws1.merge_cells('A1:E1')
        ws1.cell(1, 1).value = 'Household List'
        ws1.cell(1, 1).font = Font(bold=True, size=12)
        ws1.append([])
        ws1.append(['Household Code', 'Address', 'Purok', 'Status', 'Notes'])
        _style_headers(ws1, 3)
        for h in households:
            ws1.append([h.household_code, h.address, h.purok, h.get_status_display(), h.notes])
        _auto_width(ws1)

        # Sheet 2 — Families
        ws2 = wb.create_sheet('Families')
        ws2.merge_cells('A1:C1')
        ws2.cell(1, 1).value = 'Family List'
        ws2.cell(1, 1).font = Font(bold=True, size=12)
        ws2.append([])
        ws2.append(['Household Code', 'Family No.', 'Monthly Income Bracket'])
        _style_headers(ws2, 3)
        for h in households:
            for f in h.families.all():
                ws2.append([h.household_code, f.family_number, f.get_monthly_income_bracket_display()])
        _auto_width(ws2)

        # Sheet 3 — Members
        ws3 = wb.create_sheet('Members')
        ws3.merge_cells('A1:O1')
        ws3.cell(1, 1).value = 'Household Members'
        ws3.cell(1, 1).font = Font(bold=True, size=12)
        ws3.append([])
        ws3.append([
            'Household Code', 'Family No.', 'Full Name', 'Role', 'Age', 'Gender',
            'Civil Status', 'Sectors', 'Employment Status', 'Monthly Income',
            'Household Size', 'Dependents', 'Housing Condition', 'TUPAD Eligible', 'Contact No.',
        ])
        _style_headers(ws3, 3)
        for h in households:
            for f in h.families.all():
                for m in f.members.all():
                    ws3.append([
                        h.household_code,
                        f.family_number,
                        m.full_name,
                        m.get_role_display(),
                        m.age,
                        m.get_gender_display(),
                        m.get_civil_status_display(),
                        ', '.join(m.sectors) if m.sectors else '',
                        m.get_employment_status_display(),
                        float(m.monthly_income),
                        m.household_size,
                        m.num_dependents,
                        m.get_housing_condition_display(),
                        'Yes' if m.is_tupad_eligible else 'No',
                        m.contact_number,
                    ])
        _auto_width(ws3)

        return _excel_response(wb, 'household_report.xlsx')
